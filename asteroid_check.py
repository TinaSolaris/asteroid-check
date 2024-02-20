import argparse
import sys
import os
import csv
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill


def read_args() -> tuple:
    parser = argparse.ArgumentParser()
    # nargs='*' is used because otherwise only the first letter (symbol) of the filename will be read
    # later on it will be necessary to unwrap the list created by nargs or take its first and only element
    parser.add_argument('filename', help='Reads the file under the specified filename and uses its content to create a report.', type=str, nargs='*')
    parser.add_argument('-o', '--option',
                        help='Creates a report file under the specified filename, the report includes: the statistical results,\n' +
                        'total number of asteroids, the lists of potentially hazardous asteroids (with Earth minimum orbit intersection distances), and near to Earth asteroids.', type=str, nargs='*')
    args = parser.parse_args()

    file_extension = os.path.splitext(args.filename[0])[1]
    if not file_extension == '.csv':
        print(f'The provided dataset file \'{args.filename}\' is not of a required \'*.csv\' format.')
        sys.exit()

    if not args.option:
        return tuple(args.filename)
    else:
        optional_file_extension = os.path.splitext(args.option[0])[1]
        if not optional_file_extension == '.xlsx':
            print(f'The provided report file \'{args.option}\' is not of a required \'*.xlsx\' format.')
            sys.exit()
        return_tuple = (args.filename, args.option)
        return return_tuple


def fill_in(data):
    data_reader = csv.DictReader(data)
    database = []
    for row in data_reader:
        row_tuple = (
            row['full_name'],
            row['neo'],         # indicates if the object is near Earth
            row['pha'],         # indicates if the asteroid is potentially hazardous
            row['diameter'],    # measured in km
            row['albedo'],      # a measurement of the amount of light reflected from the surface of a celestial object
            row['q'],           # perihelion - an orbit's closest point to the Sun; measured in au
            row['moid']         # Earth Minimum Orbit Intersection Distance; measured in au
        )
        database.append(row_tuple)
    return database


def read_file(args) -> list:
    fname = args[0]

    try:
        if len(args) > 1:
            with open(*fname, 'r', encoding='utf-8') as data_set:
                database = fill_in(data_set)
        else:
            with open(fname, 'r', encoding='utf-8') as data_set:
                database = fill_in(data_set)

    except FileNotFoundError:
        print(f'The dataset file named \"{fname}\" does not exist!')
        sys.exit()
    else:
        return database


def do_operations(database) -> dict:
    results = {}
    near_earth_names = []
    potentially_hazardous_names = []
    pot_haz_distances_to_earth_orbit = []
    diameters = []
    albedos = []
    perihelions = []
    near_earth_total = 0
    potentially_hazardous_total = 0

    for row in database:
        if row[1] == 'Y':
            near_earth_total += 1
            near_earth_names.append(row[0])
        if row[2] == 'Y':
            potentially_hazardous_total += 1
            potentially_hazardous_names.append(row[0])
            pot_haz_distances_to_earth_orbit.append(float(row[6]))
        if not row[3] == '':
            diameters.append(float(row[3]))
        if not row[4] == '':
            albedos.append(float(row[4]))
        if not row[5] == '':
            perihelions.append(float(row[5]))

    diameter_median = statistics.median(diameters)
    diameter_avg = sum(diameters) / len(diameters)
    albedo_avg = sum(albedos) / len(albedos)
    perihelion_median = statistics.median(perihelions)
    items_total = len(database)

    results['diameter_median'] = diameter_median
    results['diameter_average'] = diameter_avg
    results['albedo_average'] = albedo_avg
    results['perihelion_median'] = perihelion_median
    results['neo_names'] = near_earth_names
    results['neo_total'] = near_earth_total
    results['pha_names'] = potentially_hazardous_names
    results['pha_moid_in_au'] = pot_haz_distances_to_earth_orbit
    results['pha_total'] = potentially_hazardous_total
    results['items_total'] = items_total

    return results


def fill_in_custom_cell(worksheet, col, r, c, v):
    header_font = Font(color='000000', bold=True)
    header_fill = PatternFill(start_color=col, fill_type='solid')

    cell = worksheet.cell(row=r, column=c, value=v)
    cell.font = header_font
    cell.fill = header_fill


def fill_in_custom_column(worksheet, col, r, c, data_list):
    column_fill = PatternFill(start_color=col, fill_type='solid')

    for i in range(len(data_list)):
        cell = worksheet.cell(row=r+i, column=c, value=data_list[i])
        cell.fill = column_fill


def create_excel_report(data: dict, fname):
    file_name = fname[0]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    fill_in_custom_cell(worksheet, '3DACFF', 1, 1, 'diameter_median')
    worksheet.cell(row=2, column=1, value=data.get('diameter_median'))

    fill_in_custom_cell(worksheet, '3DACFF', 1, 2, 'diameter_average')
    worksheet.cell(row=2, column=2, value=data.get('diameter_average'))

    fill_in_custom_cell(worksheet, '3DACFF', 1, 3, 'albedo_average')
    worksheet.cell(row=2, column=3, value=data.get('albedo_average'))

    fill_in_custom_cell(worksheet, '3DACFF', 1, 4, 'perihelion_median')
    worksheet.cell(row=2, column=4, value=data.get('perihelion_median'))

    font1 = Font(color='FF9F3D', bold=True)
    fill_in_custom_cell(worksheet, '3DACFF', 1, 5, 'neo_total')
    neo_val_cell = worksheet.cell(row=2, column=5, value=data.get('neo_total'))
    neo_val_cell.font = font1

    font2 = Font(color='FF0000', bold=True)
    fill_in_custom_cell(worksheet, '3DACFF', 1, 6, 'pha_total')
    pha_val_cell = worksheet.cell(row=2, column=6, value=data.get('pha_total'))
    pha_val_cell.font = font2

    fill_in_custom_cell(worksheet, '3DACFF', 1, 7, 'items_total')
    worksheet.cell(row=2, column=7, value=data.get('items_total'))

    fill_in_custom_cell(worksheet, 'FF9F3D', 4, 1, 'neo_names')
    fill_in_custom_column(worksheet, 'FFD2A5', 5, 1, data.get('neo_names'))

    fill_in_custom_cell(worksheet, 'FF0000', 4, 3, 'pha_names')
    fill_in_custom_column(worksheet, 'FFA5A5', 5, 3, data.get('pha_names'))

    fill_in_custom_cell(worksheet, 'FF0000', 4, 4, 'pha_moid')
    fill_in_custom_column(worksheet, 'FFA5A5', 5, 4, data.get('pha_moid_in_au'))

    workbook.save(file_name)
    print(f'The Excel report called \'{file_name}\' has been created.')


def print_summary(data: dict):
    print('Summary:')
    print('Total number of asteroids in the dataset: ' + str(data.get('items_total')))
    print('Total number of near to Earth asteroids in the dataset: ' + str(data.get('neo_total')))
    print('Total number of potentially hazardous asteroids in the dataset: ' + str(data.get('pha_total')))


def run():
    tuple = read_args()
    data = read_file(tuple)
    report_data = do_operations(data)
    if len(tuple) == 1:
        print_summary(report_data)
    elif len(tuple) == 2:
        create_excel_report(report_data, tuple[1])


if __name__ == "__main__":
    run()
